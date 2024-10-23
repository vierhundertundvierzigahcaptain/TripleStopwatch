from time import sleep
from threading import Thread
from winsound import PlaySound, SND_ASYNC
from tkinter import *
from tkinter import font
from openpyxl import load_workbook


# Переменные
fn = 'C:\\db\\db.xlsx'
wb = load_workbook(fn)
ws = wb['data']
A1 = ws['A1']
B1 = ws['B1']
A2 = ws['A2']
B2 = ws['B2']
A3 = ws['A3']
B3 = ws['B3']
count1_1 = A1.value
count1_2 = B1.value
count2_1 = A2.value
count2_2 = B2.value
count3_1 = A3.value
count3_2 = B3.value


# Функции


def thread(func):  # Декоратор разделения функций на потоки
    def wrapper(*args, **kwargs):
        current_thread = Thread(
            target=func, args=args, kwargs=kwargs, daemon=True)
        current_thread.start()
    return wrapper


def reset1_1():  # Сброс кол-ва изделий
    global count1_1
    count1_1 = 0
    labelCount1_1['text'] = count1_1
    ws['A1'] = count1_1
    wb.save(fn)


def reset1_2():
    global count1_2
    count1_2 = 0
    labelCount1_2['text'] = count1_2
    ws['B1'] = count1_2
    wb.save(fn)


def reset2_1():
    global count2_1
    count2_1 = 0
    labelCount2_1['text'] = count2_1
    ws['A2'] = count2_1
    wb.save(fn)


def reset2_2():
    global count2_2
    count2_2 = 0
    labelCount2_2['text'] = count2_2
    ws['B2'] = count2_2
    wb.save(fn)


def reset3_1():
    global count3_1
    count3_1 = 0
    labelCount3_1['text'] = count3_1
    ws['A3'] = count3_1
    wb.save(fn)


def reset3_2():
    global count3_2
    count3_2 = 0
    labelCount3_2['text'] = count3_2
    ws['B3'] = count3_2
    wb.save(fn)


@thread  # Функция таймера
def start1():
    duration1 = 0
    if (quantityEntry1.get()).isdigit() == True:
        duration1 = int(quantityEntry1.get())

        level = rb1.get()
        if level == 1:
            global count1_1
            count1_1 += 1
            labelCount1_1['text'] = count1_1
            labelCount1_1.update()
            ws['A1'] = count1_1
            wb.save(fn)
        elif level == 2:
            global count1_2
            count1_2 += 1
            labelCount1_2['text'] = count1_2
            labelCount1_2.update()
            ws['B1'] = count1_2
            wb.save(fn)

        while duration1 >= 0:

            if duration1 > 0:
                timerLabel1['foreground'] = '#1f1'

            timerLabel1['text'] = duration1
            timerLabel1.update()
            if duration1 <= 0:
                timerLabel1['foreground'] = '#f11'
                PlaySound('C:\\db\\sound.wav',
                          SND_ASYNC)
            sleep(1)
            duration1 -= 1


@thread
def start2():
    duration2 = 0
    if (quantityEntry2.get()).isdigit() == True:
        duration2 = int(quantityEntry2.get())

        level = rb2.get()
        if level == 1:
            global count2_1
            count2_1 += 1
            labelCount2_1['text'] = count2_1
            labelCount2_1.update()
            ws['A2'] = count2_1
            wb.save(fn)
        elif level == 2:
            global count2_2
            count2_2 += 1
            labelCount2_2['text'] = count2_2
            labelCount2_2.update()
            ws['B2'] = count2_2
            wb.save(fn)

        while duration2 >= 0:

            if duration2 > 0:
                timerLabel2['foreground'] = '#1f1'

            timerLabel2['text'] = duration2
            timerLabel2.update()
            if duration2 <= 0:
                timerLabel2['foreground'] = '#f11'
                PlaySound('C:\\db\\sound.wav',
                          SND_ASYNC)
            sleep(1)
            duration2 -= 1


@thread
def start3():
    duration3 = 0
    if (quantityEntry3.get()).isdigit() == True:
        duration3 = int(quantityEntry3.get())

        level = rb3.get()
        if level == 1:
            global count3_1
            count3_1 += 1
            labelCount3_1['text'] = count3_1
            labelCount3_1.update()
            ws['A3'] = count3_1
            wb.save(fn)
        elif level == 2:
            global count3_2
            count3_2 += 1
            labelCount3_2['text'] = count3_2
            labelCount3_2.update()
            ws['B3'] = count3_2
            wb.save(fn)

        while duration3 >= 0:

            if duration3 > 0:
                timerLabel3['foreground'] = '#1f1'

            timerLabel3['text'] = duration3
            timerLabel3.update()
            if duration3 <= 0:
                timerLabel3['foreground'] = '#f11'
                PlaySound('C:\\db\\sound.wav',
                          SND_ASYNC)
            sleep(1)
            duration3 -= 1


@thread  # Функция таймера
def keyStart1(event):
    if event.keysym == 'Left':
        duration1 = 0
        if (quantityEntry1.get()).isdigit() == True:
            duration1 = int(quantityEntry1.get())

            level = rb1.get()
            if level == 1:
                global count1_1
                count1_1 += 1
                labelCount1_1['text'] = count1_1
                labelCount1_1.update()
                ws['A1'] = count1_1
                wb.save(fn)
            elif level == 2:
                global count1_2
                count1_2 += 1
                labelCount1_2['text'] = count1_2
                labelCount1_2.update()
                ws['B1'] = count1_2
                wb.save(fn)

            while duration1 >= 0:

                if duration1 > 0:
                    timerLabel1['foreground'] = '#1f1'

                timerLabel1['text'] = duration1
                timerLabel1.update()
                if duration1 <= 0:
                    timerLabel1['foreground'] = '#f11'
                    PlaySound('C:\\db\\sound.wav',
                              SND_ASYNC)
                sleep(1)
                duration1 -= 1


@thread
def keyStart2(event):
    if event.keysym == 'Down':
        duration2 = 0
        if (quantityEntry2.get()).isdigit() == True:
            duration2 = int(quantityEntry2.get())

            level = rb2.get()
            if level == 1:
                global count2_1
                count2_1 += 1
                labelCount2_1['text'] = count2_1
                labelCount2_1.update()
                ws['A2'] = count2_1
                wb.save(fn)
            elif level == 2:
                global count2_2
                count2_2 += 1
                labelCount2_2['text'] = count2_2
                labelCount2_2.update()
                ws['B2'] = count2_2
                wb.save(fn)

            while duration2 >= 0:

                if duration2 > 0:
                    timerLabel2['foreground'] = '#1f1'

                timerLabel2['text'] = duration2
                timerLabel2.update()
                if duration2 <= 0:
                    timerLabel2['foreground'] = '#f11'
                    PlaySound('C:\\db\\sound.wav',
                              SND_ASYNC)
                sleep(1)
                duration2 -= 1


@thread
def keyStart3(event):
    if event.keysym == 'Right':
        duration3 = 0
        if (quantityEntry3.get()).isdigit() == True:
            duration3 = int(quantityEntry3.get())

            level = rb3.get()
            if level == 1:
                global count3_1
                count3_1 += 1
                labelCount3_1['text'] = count3_1
                labelCount3_1.update()
                ws['A3'] = count3_1
                wb.save(fn)
            elif level == 2:
                global count3_2
                count3_2 += 1
                labelCount3_2['text'] = count3_2
                labelCount3_2.update()
                ws['B3'] = count3_2
                wb.save(fn)

            while duration3 >= 0:

                if duration3 > 0:
                    timerLabel3['foreground'] = '#1f1'

                timerLabel3['text'] = duration3
                timerLabel3.update()
                if duration3 <= 0:
                    timerLabel3['foreground'] = '#f11'
                    PlaySound('C:\\db\\sound.wav',
                              SND_ASYNC)
                sleep(1)
                duration3 -= 1


# Создание окна
root = Tk()

root['bg'] = '#111'
root.title('Тройной секундомер')
root.geometry('1200x1000')

root.resizable(width=False, height=False)

canvas = Canvas(root, width=1200, height=1000)
canvas.pack()

# Разметка окна
frame1 = Frame(root, bg='#111', width=400, height=1000)
frame2 = Frame(root, bg='#111', width=400, height=1000)
frame3 = Frame(root, bg='#111', width=400, height=1000)
frame1.place(relwidth=0.33, relheight=1)
frame2.place(relx=0.33, relwidth=0.34, relheight=1)
frame3.place(relx=0.67, relwidth=0.33, relheight=1)

# Наполнение первой области
font1 = font.Font(family='Arial', size=16, weight='bold')
font2 = font.Font(family='Arial', size=150, weight='bold')

title1 = Label(frame1, text='Кол-во изделий:',
               background='#111', foreground='#fff', font=font1)
title1.place(relx=0.5, rely=0.025, anchor=N)

rb1 = IntVar()

radiobutton1_1 = Radiobutton(
    frame1, text='', variable=rb1, value=1, background='#111')
radiobutton1_1.place(relx=0.26, rely=0.06, anchor=N)

radiobutton1_2 = Radiobutton(
    frame1, text='', variable=rb1, value=2, background='#111')
radiobutton1_2.place(relx=0.26, rely=0.1, anchor=N)

labelCount1_1 = Label(frame1, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount1_1.place(relx=0.5, rely=0.06, anchor=N)

labelCount1_2 = Label(frame1, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount1_2.place(relx=0.5, rely=0.1, anchor=N)

btnReset1_1 = Button(frame1, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset1_1)
btnReset1_1.place(relx=0.75, rely=0.055, anchor=N)

btnReset1_2 = Button(frame1, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset1_2)
btnReset1_2.place(relx=0.75, rely=0.1, anchor=N)

timerLabel1 = Label(frame1, text='000',
                    background='#111', foreground='#1f1', font=font2)
timerLabel1.place(relx=0.5, rely=0.4, anchor=N)

quantityLabel1 = Label(frame1, text='Введите время:',
                       background='#111', foreground='#fff', font=font1)
quantityLabel1.place(relx=0.5, rely=0.8, anchor=N)

quantityEntry1 = Entry(frame1, width=5, text='1', bg="white", font=font1)
quantityEntry1.place(relx=0.5, rely=0.85, anchor=N)

startBtn1 = Button(frame1, text='Нажать ← для старта', background='#111',
                   foreground='#fff', font=font1, command=start1)
startBtn1.place(relx=0.5, rely=0.9, anchor=N)
canvas.bind_all("<KeyPress-Left>", keyStart1)

# Наполнение второй области
title2 = Label(frame2, text='Кол-во изделий:',
               background='#111', foreground='#fff', font=font1)
title2.place(relx=0.5, rely=0.025, anchor=N)

rb2 = IntVar()

radiobutton2_1 = Radiobutton(
    frame2, text='', variable=rb2, value=1, background='#111')
radiobutton2_1.place(relx=0.26, rely=0.06, anchor=N)

radiobutton2_2 = Radiobutton(
    frame2, text='', variable=rb2, value=2, background='#111')
radiobutton2_2.place(relx=0.26, rely=0.1, anchor=N)

labelCount2_1 = Label(frame2, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount2_1.place(relx=0.5, rely=0.06, anchor=N)

labelCount2_2 = Label(frame2, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount2_2.place(relx=0.5, rely=0.1, anchor=N)

btnReset2_1 = Button(frame2, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset2_1)
btnReset2_1.place(relx=0.75, rely=0.055, anchor=N)

btnReset2_2 = Button(frame2, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset2_2)
btnReset2_2.place(relx=0.75, rely=0.1, anchor=N)

timerLabel2 = Label(frame2, text='000',
                    background='#111', foreground='#1f1', font=font2)
timerLabel2.place(relx=0.5, rely=0.4, anchor=N)

quantityLabel2 = Label(frame2, text='Введите время:',
                       background='#111', foreground='#fff', font=font1)
quantityLabel2.place(relx=0.5, rely=0.8, anchor=N)

quantityEntry2 = Entry(frame2, width=5, text='2', bg="white", font=font1)
quantityEntry2.place(relx=0.5, rely=0.85, anchor=N)

startBtn2 = Button(frame2, text='Нажать ↓ для старта', background='#111',
                   foreground='#fff', font=font1, command=start2)
startBtn2.place(relx=0.5, rely=0.9, anchor=N)
canvas.bind_all("<KeyPress-Down>", keyStart2)

# Наполнение трейтьей области
title3 = Label(frame3, text='Кол-во изделий:',
               background='#111', foreground='#fff', font=font1)
title3.place(relx=0.5, rely=0.025, anchor=N)

rb3 = IntVar()

radiobutton3_1 = Radiobutton(
    frame3, text='', variable=rb3, value=1, background='#111')
radiobutton3_1.place(relx=0.26, rely=0.06, anchor=N)

radiobutton3_2 = Radiobutton(
    frame3, text='', variable=rb3, value=2, background='#111')
radiobutton3_2.place(relx=0.26, rely=0.1, anchor=N)

labelCount3_1 = Label(frame3, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount3_1.place(relx=0.5, rely=0.06, anchor=N)

labelCount3_2 = Label(frame3, text='0',
                      background='#111', foreground='#fff', font=font1)
labelCount3_2.place(relx=0.5, rely=0.1, anchor=N)

btnReset3_1 = Button(frame3, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset3_1)
btnReset3_1.place(relx=0.75, rely=0.055, anchor=N)

btnReset3_2 = Button(frame3, text='Сброс', background='#111',
                     foreground='#fff', font=font1, command=reset3_2)
btnReset3_2.place(relx=0.75, rely=0.1, anchor=N)

timerLabel3 = Label(frame3, text='000',
                    background='#111', foreground='#1f1', font=font2)
timerLabel3.place(relx=0.5, rely=0.4, anchor=N)

quantityLabel3 = Label(frame3, text='Введите время:',
                       background='#111', foreground='#fff', font=font1)
quantityLabel3.place(relx=0.5, rely=0.8, anchor=N)

quantityEntry3 = Entry(frame3, width=5, text='3', bg="white", font=font1)
quantityEntry3.place(relx=0.5, rely=0.85, anchor=N)

startBtn3 = Button(frame3, text='Нажать → для старта', background='#111',
                   foreground='#fff', font=font1, command=start3)
startBtn3.place(relx=0.5, rely=0.9, anchor=N)
canvas.bind_all("<KeyPress-Right>", keyStart3)

# Присвоение счётчикам значений при открытии программы
labelCount1_1['text'] = count1_1
labelCount1_1.update()
labelCount1_2['text'] = count1_2
labelCount1_2.update()
labelCount2_1['text'] = count2_1
labelCount2_1.update()
labelCount2_2['text'] = count2_2
labelCount2_2.update()
labelCount3_1['text'] = count3_1
labelCount3_1.update()
labelCount3_2['text'] = count3_2
labelCount3_2.update()

#   Бесконечный цикл для работы окна
root.mainloop()

# Компиляция
# pyinstaller --onefile --icon=icon.ico tripleStopwatch.py
# python -m nuitka --follow-imports --windows-icon-from-ico=icon.ico tripleStopwatch.py
